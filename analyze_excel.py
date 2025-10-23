import openpyxl

wb = openpyxl.load_workbook('standart.xlsx')
ws = wb.active

print('Sheet adları:', wb.sheetnames)
print('\nKolon başlıkları:')
headers = [cell.value for cell in ws[1]]
print(headers)

print('\nİlk 10 satır:')
for i in range(2, min(12, ws.max_row+1)):
    row_data = [cell.value for cell in ws[i]]
    print(f"Satır {i-1}: {row_data}")

print(f'\nToplam satır sayısı: {ws.max_row-1}')
print(f'Toplam kolon sayısı: {ws.max_column}')
