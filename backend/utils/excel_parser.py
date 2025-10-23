import openpyxl
from typing import List, Dict, Optional
import logging

from models.profile import Profile

logger = logging.getLogger(__name__)


# Excel'deki kategori kolonları ve ölçü mapping'leri
CATEGORY_COLUMNS = {
    "STANDART BORU": {
        "code_col": 0,
        "dimensions": {"Ø": 1, "K": 2}
    },
    "STANDART KUTU": {
        "code_col": 4,
        "dimensions": {"A": 5, "B": 6, "K": 7}
    },
    "STANDART T": {
        "code_col": 8,
        "dimensions": {"A": 9, "B": 10, "K": 11}
    },
    "STANDART U": {
        "code_col": 12,
        "dimensions": {"A": 13, "B": 14, "K": 15}
    },
    "STANDART LAMA": {
        "code_col": 16,
        "dimensions": {"A": 17, "B": 18}
    },
    "STANDART KÖŞEBENT": {
        "code_col": 20,
        "dimensions": {"A": 21, "B": 22, "K": 23}
    }
}


def parse_excel_file(file_path: str) -> List[Profile]:
    """
    Excel dosyasını parse eder ve Profile listesi döner
    
    Args:
        file_path: Excel dosyasının yolu
        
    Returns:
        Profile nesnelerinin listesi
    """
    logger.info(f"Excel dosyası parse ediliyor: {file_path}")
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        
        profiles = []
        
        # Her satırı işle (başlık satırını atla, 2. satırdan başla)
        for row_idx in range(2, worksheet.max_row + 1):
            row_profiles = _parse_row(worksheet, row_idx)
            profiles.extend(row_profiles)
        
        logger.info(f"Toplam {len(profiles)} profil parse edildi")
        return profiles
        
    except Exception as e:
        logger.error(f"Excel parse hatası: {e}")
        raise


def _parse_row(worksheet, row_idx: int) -> List[Profile]:
    """
    Tek bir satırı parse eder ve tüm kategorilerdeki profilleri döner
    
    Args:
        worksheet: Excel worksheet
        row_idx: Satır numarası
        
    Returns:
        Bu satırdaki Profile nesnelerinin listesi
    """
    profiles = []
    
    # Her kategori için kontrol et
    for category_name, config in CATEGORY_COLUMNS.items():
        code_col = config["code_col"]
        dimension_cols = config["dimensions"]
        
        # Profil kodunu al
        code_cell = worksheet.cell(row=row_idx, column=code_col + 1)
        code = code_cell.value
        
        # Kod yoksa veya boşsa bu kategoriyi atla
        if not code or not isinstance(code, str):
            continue
        
        # Ölçüleri al
        dimensions = {}
        has_valid_dimension = False
        
        for dim_name, dim_col in dimension_cols.items():
            dim_cell = worksheet.cell(row=row_idx, column=dim_col + 1)
            dim_value = dim_cell.value
            
            # Ölçü değeri varsa ve sayısal ise ekle
            if dim_value is not None and isinstance(dim_value, (int, float)):
                dimensions[dim_name] = float(dim_value)
                has_valid_dimension = True
        
        # En az bir geçerli ölçü varsa profili ekle
        if has_valid_dimension:
            text_repr = _create_text_representation(code, category_name, dimensions)
            
            profile = Profile(
                code=code,
                category=category_name,
                dimensions=dimensions,
                text_representation=text_repr
            )
            
            profiles.append(profile)
    
    return profiles


def _create_text_representation(code: str, category: str, dimensions: Dict[str, float]) -> str:
    """
    Profil için text representation oluşturur
    
    Args:
        code: Profil kodu
        category: Kategori adı
        dimensions: Ölçüler dictionary
        
    Returns:
        Text formatında profil bilgisi
    """
    dims_str = ", ".join([f"{k}={v}mm" for k, v in dimensions.items()])
    
    return (
        f"Profil Kodu: {code}\n"
        f"Kategori: {category}\n"
        f"Ölçüler: {dims_str}"
    )


def validate_profiles(profiles: List[Profile]) -> List[Profile]:
    """
    Profil listesini validate eder ve geçersiz olanları filtreler
    
    Args:
        profiles: Profile listesi
        
    Returns:
        Validate edilmiş Profile listesi
    """
    valid_profiles = []
    
    for profile in profiles:
        # Kod kontrolü
        if not profile.code or len(profile.code.strip()) == 0:
            logger.warning(f"Geçersiz profil kodu atlandı: {profile.code}")
            continue
        
        # Ölçü kontrolü
        if not profile.dimensions or len(profile.dimensions) == 0:
            logger.warning(f"Ölçüsüz profil atlandı: {profile.code}")
            continue
        
        # Tüm ölçülerin pozitif olduğunu kontrol et
        if any(v <= 0 for v in profile.dimensions.values()):
            logger.warning(f"Negatif veya sıfır ölçülü profil atlandı: {profile.code}")
            continue
        
        valid_profiles.append(profile)
    
    logger.info(f"{len(valid_profiles)}/{len(profiles)} profil validate edildi")
    return valid_profiles


def get_category_summary(profiles: List[Profile]) -> Dict[str, int]:
    """
    Kategorilere göre profil sayılarını döner
    
    Args:
        profiles: Profile listesi
        
    Returns:
        Kategori adı -> profil sayısı dictionary
    """
    summary = {}
    
    for profile in profiles:
        category = profile.category
        summary[category] = summary.get(category, 0) + 1
    
    return summary
